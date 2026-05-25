from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from app.gmail_client import get_gmail_service
from app.gmail_helpers import extract_headers, get_message_body
from app.parsers import parse_job

ROOT = Path(__file__).resolve().parent.parent
JOBS_MASTER = ROOT / "jobs_master.xlsx"


def load_existing_keys(ws, email_id_col=8, link_col=6):
    """
    Dedupe keys:
      - email_id (Gmail message id)
      - link (job link)
    """
    email_ids = set()
    links = set()

    for r in range(2, ws.max_row + 1):
        email_id = ws.cell(r, email_id_col).value
        link = ws.cell(r, link_col).value

        if email_id:
            email_ids.add(str(email_id).strip())
        if link:
            links.add(str(link).strip())

    return email_ids, links

def ensure_jobs_sheet(wb):
    """
    Ensure a worksheet named 'Jobs' exists.
    If created, also write header row.
    Returns the worksheet.
    """
    if "Jobs" in wb.sheetnames:
        ws = wb["Jobs"]
    else:
        ws = wb.create_sheet("Jobs")
        ws.append([
            "received_at",
            "source",
            "title",
            "company",
            "location",
            "link",
            "description",
            "email_id",
            "contact_email",
        ])

        # Optional: remove default blank sheet if present
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

    # ✅ Ensure header includes contact_email (upgrade old masters safely)
    if ws.max_row >= 1:
        existing_header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        existing_header_norm = [str(x).strip().lower() if x else "" for x in existing_header]

        if "contact_email" not in existing_header_norm:
            ws.cell(row=1, column=len(existing_header) + 1, value="contact_email")

    return ws


def main():
    svc = get_gmail_service()

    q = 'newer_than:7d -in:sent -from:mailer-daemon -from:postmaster (from:linkedin OR from:indeed OR from:rozee OR from:freelancer)'
    res = svc.users().messages().list(userId="me", q=q, maxResults=100).execute()
    msgs = res.get("messages", [])

    if not JOBS_MASTER.exists():
        raise FileNotFoundError(f"{JOBS_MASTER} not found.")

    wb = load_workbook(JOBS_MASTER)
    ws = ensure_jobs_sheet(wb)

    # Determine column indexes from header (so upgrades don't break)
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_norm = [str(x).strip().lower() if x else "" for x in header]

    def col_idx(name: str, default: int) -> int:
        name = name.lower()
        if name in header_norm:
            return header_norm.index(name) + 1
        return default

    EMAIL_ID_COL = col_idx("email_id", 8)
    LINK_COL = col_idx("link", 6)
    CONTACT_COL = col_idx("contact_email", 9)

    existing_email_ids, existing_links = load_existing_keys(
        ws,
        email_id_col=EMAIL_ID_COL,
        link_col=LINK_COL
    )

    added = 0

    for m in msgs:
        msg = svc.users().messages().get(
            userId="me",
            id=m["id"],
            format="full"
        ).execute()

        email_id = msg.get("id")
        if not email_id:
            continue

        email_id_s = str(email_id).strip()
        if email_id_s in existing_email_ids:
            continue

        headers = extract_headers(msg["payload"])
        subject = headers.get("subject", "") or ""
        from_email = headers.get("from", "") or ""
        body = get_message_body(msg["payload"]) or ""

        # ✅ Only call parse_job once
        job = parse_job(subject, body, from_email=from_email) or {}

         # ✅ Skip non-job emails
        if not job.get("is_job"):
            continue

        link = job.get("link")
        contact_email = job.get("contact_email")

        # ✅ Dedupe by link
        if link and str(link).strip() in existing_links:
            continue

        received_at = datetime.fromtimestamp(
            int(msg["internalDate"]) / 1000
        ).isoformat(timespec="seconds")

        # detect source
        src = "Other"
        f = from_email.lower()
        if "linkedin" in f:
            src = "LinkedIn"
        elif "indeed" in f:
            src = "Indeed"
        elif "glassdoor" in f:
            src = "Glassdoor"

        # ✅ Build row according to header order
        row_map = {
            "received_at": received_at,
            "source": src,
            "title": job.get("title"),
            "company": job.get("company"),
            "location": job.get("location"),
            "link": link,
            "description": job.get("description"),
            "email_id": email_id_s,
            "contact_email": contact_email,
        }

        # Ensure we write all columns present in the header
        out_row = []
        for h in header_norm:
            if h == "":
                out_row.append("")
            else:
                out_row.append(row_map.get(h))

        # If header grew after we cached it, include contact_email
        if "contact_email" not in header_norm:
            out_row.append(contact_email)

        ws.append(out_row)

        # Make link clickable (find actual link column)
        if link:
            row_idx = ws.max_row
            link_cell = ws.cell(row=row_idx, column=LINK_COL)
            link_cell.hyperlink = str(link).strip()
            link_cell.style = "Hyperlink"

        existing_email_ids.add(email_id_s)
        if link:
            existing_links.add(str(link).strip())

        added += 1

    wb.save(JOBS_MASTER)
    print(f"Jobs added to jobs_master.xlsx: {added}")


if __name__ == "__main__":
    main()