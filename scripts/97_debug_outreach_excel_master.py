from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
FILE = ROOT / "exports" / "outreach_master.xlsx"

def header_map(ws):
    hm = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hm[str(v).strip()] = c
    return hm

def get(ws, r, hm, key):
    c = hm.get(key)
    return ws.cell(r, c).value if c else None

def main():
    wb = load_workbook(FILE)
    print("File:", FILE)
    print("Sheets:", wb.sheetnames)
    ws = wb["Outreach"]
    hm = header_map(ws)

    print("\nHeaders:", sorted(hm.keys()))
    print("Rows:", ws.max_row - 1)

    for r in range(2, min(ws.max_row, 10) + 1):
        to_email = get(ws, r, hm, "to_email")
        send_flag = str(get(ws, r, hm, "send_flag") or "").strip().upper()
        status = str(get(ws, r, hm, "status") or "").strip().upper()
        subject = get(ws, r, hm, "subject")

        print(f"Row {r}: email={to_email} send_flag={send_flag} status={status} subject={subject}")

if __name__ == "__main__":
    main()
