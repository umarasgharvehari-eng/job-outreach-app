from pathlib import Path
from openpyxl import load_workbook
from app.sheets_client import get_sheets_service

SHEET_ID = "1DXRX_Q5B9EbryWMyMMHgGfxeOoF8ytu_6KLyMX-z02I"

ROOT = Path(__file__).resolve().parent.parent
OUTREACH_MASTER = ROOT / "exports" / "outreach_master.xlsx"

EXCEL_HEADERS = [
    "to_email","to_name","subject","message_template","status",
    "followup_due_at","followup_sent_at","thread_id",
    "person_id","send_flag","sent_at","batch_id",
    "job_title","company","job_link","last_reply_at","last_reply_snippet"
]

def main():
    sheets = get_sheets_service()

    data = sheets.spreadsheets().values().get(
        spreadsheetId=SHEET_ID,
        range="Outreach!A1:H5000"
    ).execute().get("values", [])

    if not data or len(data) < 2:
        print("No outreach rows found in Google Sheet.")
        return

    header = data[0]
    rows = data[1:]

    idx = {h.strip(): i for i, h in enumerate(header) if h}

    # required in sheet
    for k in ["to_email", "to_name", "subject", "status", "thread_id", "sent_at"]:
        if k not in idx:
            print("Missing in Sheet header:", k)
            print("Sheet header is:", header)
            return

    wb = load_workbook(OUTREACH_MASTER)
    ws = wb["Outreach"]

    # rewrite header (ensures perfect structure)
    ws.delete_rows(1, ws.max_row)
    ws.append(EXCEL_HEADERS)

    exported = 0
    for r in rows:
        to_email = r[idx["to_email"]] if idx["to_email"] < len(r) else ""
        to_name = r[idx["to_name"]] if idx["to_name"] < len(r) else ""
        subject = r[idx["subject"]] if idx["subject"] < len(r) else ""
        status = r[idx["status"]] if idx["status"] < len(r) else ""
        sent_at = r[idx["sent_at"]] if idx["sent_at"] < len(r) else ""
        thread_id = r[idx["thread_id"]] if idx["thread_id"] < len(r) else ""
        followup_due_at = r[idx["followup_due_at"]] if idx["followup_due_at"] < len(r) else ""
        followup_sent_at = r[idx["followup_sent_at"]] if idx["followup_sent_at"] < len(r) else ""

        # send_flag rule
        send_flag = "YES" if str(status).strip().upper() in ("", "NEW") else ""

        out = [
            to_email, to_name, subject, "", status,
            followup_due_at, followup_sent_at, thread_id,
            "", send_flag, sent_at, "",
            "", "", "", "", ""
        ]
        ws.append(out)
        exported += 1

    wb.save(OUTREACH_MASTER)
    print(f"Exported {exported} rows -> {OUTREACH_MASTER}")

if __name__ == "__main__":
    main()
