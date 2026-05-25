from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
EXPORTS = ROOT / "exports"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def auto_width(ws, max_width=55):
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), max_width)


def style_sheet(ws):
    # Header row styling
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # Body cells wrap
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = BODY_ALIGN

    # Freeze header + filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    auto_width(ws)


def add_outreach_dropdowns(ws):
    # find column indexes by header name
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    # send_flag dropdown YES/NO
    if "send_flag" in headers:
        col = headers["send_flag"]
        dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}1048576")

    # status dropdown
    if "status" in headers:
        col = headers["status"]
        dv = DataValidation(
            type="list",
            formula1='"NEW,QUEUED,SENT,FOLLOWUP_SENT,REPLIED,STOP"',
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}1048576")


def format_file(path: Path, sheet_name: str, outreach_dropdowns: bool = False):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    style_sheet(ws)
    if outreach_dropdowns:
        add_outreach_dropdowns(ws)
    wb.save(path)
    print("Formatted:", path.name)


def main():
    format_file(EXPORTS / "outreach_master.xlsx", "Outreach", outreach_dropdowns=True)
    format_file(EXPORTS / "jobs_master.xlsx", "Jobs")
    format_file(EXPORTS / "replies_master.xlsx", "Replies")


if __name__ == "__main__":
    main()
