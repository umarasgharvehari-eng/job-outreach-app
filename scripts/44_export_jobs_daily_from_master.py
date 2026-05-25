# scripts/44_export_jobs_daily_from_master.py
# Daily Excel exporter (from jobs_master.xlsx) + professional polish + clickable hyperlinks

from pathlib import Path
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule


ROOT = Path(__file__).resolve().parent.parent

# If your master jobs file is in a different path, change it here:
MASTER = ROOT / "exports" / "jobs_master.xlsx"

OUT_DIR = ROOT / "exports" / "daily_jobs"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def autosize(ws, max_width=60):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, max_width)


def polish(ws):
    # Header style
    header_fill = PatternFill(start_color="2F2F2F", end_color="2F2F2F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze header + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Wrap text for long fields
    wrap_cols = {"title", "subject", "description"}
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    for idx, h in enumerate(headers, start=1):
        if h in wrap_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, idx).alignment = Alignment(wrap_text=True, vertical="top")

    # Conditional formatting for match_flag if present
    if "match_flag" in headers:
        col = headers.index("match_flag") + 1
        col_letter = get_column_letter(col)
        rng = f"{col_letter}2:{col_letter}{ws.max_row}"

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        rule = FormulaRule(formula=[f'${col_letter}2="YES"'], fill=green_fill)
        ws.conditional_formatting.add(rng, rule)


def make_links_clickable(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    link_col = None

    # Your master typically uses "link"
    if "job_link" in headers:
        link_col = headers.index("job_link") + 1
    elif "link" in headers:
        link_col = headers.index("link") + 1

    if not link_col:
        return

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, link_col)
        val = str(cell.value or "").strip()
        if val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(color="0000FF", underline="single")


def main():
    today = date.today().isoformat()
    out = OUT_DIR / f"jobs_{today}.xlsx"

    if not MASTER.exists():
        print("Master file not found:", MASTER)
        return

    # Read master
    wb_m = load_workbook(MASTER)
    ws_m = wb_m.active

    # Master headers
    headers = [ws_m.cell(1, c).value for c in range(1, ws_m.max_column + 1)]
    headers = [h for h in headers if h]

    # Create/open daily file
    if out.exists():
        wb = load_workbook(out)
        ws = wb.active

        # If header mismatches, recreate file cleanly
        existing_header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        existing_header = [h for h in existing_header if h]
        if existing_header != headers:
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs"
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"
        ws.append(headers)

    # Column index in master by name
    def col_index(name):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    col_received = col_index("received_at") or col_index("sent_at")
    col_link = col_index("job_link") or col_index("link")
    col_subject = col_index("subject") or col_index("email_subject")

    # Build header index for daily file
    daily_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hidx = {h: i + 1 for i, h in enumerate(daily_headers) if h}

    # Existing dedupe keys (link + subject)
    existing = set()
    for r in range(2, ws.max_row + 1):
        link = ""
        subj = ""

        if "job_link" in hidx:
            link = (ws.cell(r, hidx["job_link"]).value or "").strip()
        elif "link" in hidx:
            link = (ws.cell(r, hidx["link"]).value or "").strip()

        if "subject" in hidx:
            subj = (ws.cell(r, hidx["subject"]).value or "").strip()
        elif "email_subject" in hidx:
            subj = (ws.cell(r, hidx["email_subject"]).value or "").strip()

        existing.add((link, subj))

    added = 0

    # Append today's rows from master (based on received_at containing today's date)
    for r in range(2, ws_m.max_row + 1):
        received = str(ws_m.cell(r, col_received).value or "") if col_received else ""

        # Robust check: keep rows that include today's date string
        if today not in received:
            continue

        link = (ws_m.cell(r, col_link).value or "").strip() if col_link else ""
        subj = (ws_m.cell(r, col_subject).value or "").strip() if col_subject else ""
        key = (link, subj)

        if key in existing:
            continue

        row_values = [ws_m.cell(r, c).value for c in range(1, len(headers) + 1)]
        ws.append(row_values)
        existing.add(key)
        added += 1

    autosize(ws)
    polish(ws)
    make_links_clickable(ws)

    wb.save(out)

    print(f"Daily jobs file: {out}")
    print(f"Jobs added today from master: {added}")


if __name__ == "__main__":
    main()
