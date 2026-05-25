from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUTREACH_MASTER = ROOT / "exports" / "outreach_master.xlsx"

def header_map(ws):
    hm = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hm[str(v).strip()] = c
    return hm

def cell(ws, r, hm, key):
    c = hm.get(key)
    return ws.cell(r, c).value if c else None

def main():
    print("File:", OUTREACH_MASTER)
    wb = load_workbook(OUTREACH_MASTER)

    print("Sheets:", wb.sheetnames)
    if "Outreach" not in wb.sheetnames:
        print("❌ Sheet name 'Outreach' not found. Your sheet name is different.")
        return

    ws = wb["Outreach"]
    hm = header_map(ws)

    print("\nHeaders found:")
    for k in sorted(hm.keys()):
        print(" -", k)

    required = ["to_email", "send_flag", "status"]
    missing = [k for k in required if k not in hm]
    if missing:
        print("\n❌ Missing required headers:", missing)
        print("Fix: rename your header cells exactly to these:", required)
        return

    eligible = 0
    total = ws.max_row - 1
    print(f"\nTotal data rows: {total}")

    # Show first 20 rows and why they are skipped
    for r in range(2, min(ws.max_row, 21) + 1):
        to_email = str(cell(ws, r, hm, "to_email") or "").strip()
        send_flag = str(cell(ws, r, hm, "send_flag") or "").strip().upper()
        status = str(cell(ws, r, hm, "status") or "").strip().upper()

        reasons = []
        if not to_email:
            reasons.append("to_email empty")
        if send_flag != "YES":
            reasons.append(f"send_flag='{send_flag}' (needs YES)")
        if status not in ("", "NEW"):
            reasons.append(f"status='{status}' (needs blank/NEW)")

        if not reasons:
            eligible += 1
            print(f"✅ Row {r} eligible -> {to_email}")
        else:
            print(f"⛔ Row {r} skipped -> {', '.join(reasons)}")

    print(f"\nEligible rows in first 20: {eligible}")
    print("If eligible=0, fix Excel values/headers and re-run send script.")

if __name__ == "__main__":
    main()
