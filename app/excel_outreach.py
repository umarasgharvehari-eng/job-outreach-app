from __future__ import annotations
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

from app.sheet_settings import load_settings
from app.gmail_helpers import send_email_with_pdf
from app.db import get_conn
from app.config import CV_PDF_PATH


def _header_map(ws) -> Dict[str, int]:
    hm = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hm[str(v).strip()] = c
    return hm


def _get(ws, r: int, hm: Dict[str, int], key: str):
    c = hm.get(key)
    return ws.cell(r, c).value if c else None


def _set(ws, r: int, hm: Dict[str, int], key: str, value):
    c = hm.get(key)
    if c:
        ws.cell(r, c).value = value


def _build_email(to_name: str, job_title: str, company: str, template: Optional[str], job_link: str):
    name = to_name or "there"
    title = job_title or "the role"
    comp = company or "your company"

    if template and str(template).strip():
        body = str(template).format(name=name, title=title, company=comp)
    else:
        body = f"""Hi {name},

I hope you’re doing well. I’m applying for {title} at {comp}.
Please find my CV attached.

Thanks,
Umar
"""

    if job_link:
        body += f"\nJob Link: {job_link}\n"

    subject = f"Application: {title} — Umar"
    return subject, body


def send_from_outreach_master(
    svc,
    outreach_path: Path,
    max_per_run: int = 10,
) -> int:
    settings = load_settings()

    # daily cap check (DB-based)
    today_str = date.today().isoformat()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute("""SELECT COUNT(*) FROM outreach WHERE date(sent_at)=?""", (today_str,))
        today_count = cur.fetchone()[0]
        if today_count >= settings["max_emails_per_day"]:
            print("Daily limit reached.")
            return 0

    wb = load_workbook(outreach_path)
    ws = wb["Outreach"]
    hm = _header_map(ws)

    now = datetime.now()
    batch_id = f"B{now.strftime('%Y%m%d_%H%M%S')}"
    sent_count = 0

    with get_conn() as conn:
        cur = conn.cursor()

        for r in range(2, ws.max_row + 1):
            if sent_count >= max_per_run:
                break

            send_flag = str(_get(ws, r, hm, "send_flag") or "").strip().upper()
            status = str(_get(ws, r, hm, "status") or "").strip().upper()
            to_email = str(_get(ws, r, hm, "to_email") or "").strip()

            if send_flag != "YES":
                continue
            if status not in ("", "NEW"):
                continue
            if not to_email:
                continue
            if status != "NEW":
                continue

            if not to_email or "@" not in str(to_email):
                continue
            to_name = str(_get(ws, r, hm, "to_name") or "").strip()
            job_title = str(_get(ws, r, hm, "job_title") or "").strip()
            company = str(_get(ws, r, hm, "company") or "").strip()
            job_link = str(_get(ws, r, hm, "job_link") or "").strip()
            template = _get(ws, r, hm, "message_template")

            subject, body = _build_email(to_name, job_title, company, template, job_link)

            # Followup due at exact time from settings
            due = now + timedelta(days=settings["followup_days"])
            followup_due_at = due.replace(
                hour=settings.get("followup_hour", 11),
                minute=settings.get("followup_minute", 0),
                second=0,
                microsecond=0
            )

            sent_message_id, thread_id = send_email_with_pdf(
                svc,
                to_email=to_email,
                subject=subject,
                body_text=body,
                pdf_path=CV_PDF_PATH,
                thread_id=None
            )

            _set(ws, r, hm, "status", "SENT")
            _set(ws, r, hm, "batch_id", batch_id)
            _set(ws, r, hm, "sent_at", now.isoformat(timespec="seconds"))
            _set(ws, r, hm, "followup_due_at", followup_due_at.isoformat(timespec="seconds"))
            _set(ws, r, hm, "thread_id", thread_id)

            cur.execute("""
                INSERT INTO outreach
                (to_email, to_name, subject, status, sent_at, followup_due_at, gmail_thread_id, gmail_sent_message_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                to_email,
                to_name,
                subject,
                "SENT",
                now.isoformat(timespec="seconds"),
                followup_due_at.isoformat(timespec="seconds"),
                thread_id,
                sent_message_id,
                datetime.now().isoformat(timespec="seconds")
            ))

            sent_count += 1

        conn.commit()

    wb.save(outreach_path)
    return sent_count
