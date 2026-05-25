from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _auto_width(ws, max_width: int = 55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is None:
                continue
            s = str(val)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), max_width)


def write_daily_xlsx_pro(
    path: Path,
    rows: List[Dict],
    columns: List[str],
    sheet_name: str = "Jobs",
    date_cols: Optional[List[str]] = None,
    hyperlink_cols: Optional[List[str]] = None,
):
    """
    Professional Excel export:
    - Ordered columns
    - Header styling + filters
    - Freeze top row
    - Wrap text
    - Auto column widths
    - Hyperlinks for URL columns
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    date_cols = date_cols or []
    hyperlink_cols = hyperlink_cols or []

    if not rows:
        # create empty sheet with headers
        df = pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(rows)
        # ensure missing columns exist
        for c in columns:
            if c not in df.columns:
                df[c] = None
        df = df[columns]

    # normalize datetime strings if present
    for dc in date_cols:
        if dc in df.columns:
            def _to_dt(x):
                if not x:
                    return None
                try:
                    # supports ISO strings
                    return datetime.fromisoformat(str(x).replace("Z", ""))
                except Exception:
                    return x
            df[dc] = df[dc].apply(_to_dt)

    # write with pandas first
    df.to_excel(path, index=False, sheet_name=sheet_name)

    # style with openpyxl
    wb = load_workbook(path)
    ws = wb[sheet_name]

    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

    body_align = Alignment(vertical="top", wrap_text=True)

    # header style
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # body style + hyperlink
    header_map = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = body_align

    # hyperlinks
    for col_name in hyperlink_cols:
        if col_name in header_map:
            cidx = header_map[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=cidx)
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    # freeze header
    ws.freeze_panes = "A2"

    # auto-filter
    ws.auto_filter.ref = ws.dimensions

    # set nice row height
    ws.row_dimensions[1].height = 22

    # date formatting
    for col_name in date_cols:
        if col_name in header_map:
            cidx = header_map[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=cidx)
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"

    _auto_width(ws)

    wb.save(path)
