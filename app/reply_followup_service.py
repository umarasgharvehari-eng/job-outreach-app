from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

from app.gmail_helpers import send_email_with_pdf
from app.config import CV_PDF_PATH


def _header_map(ws) -> Dict[str, int]:
    hm = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hm[str(v).strip()] = c
    return hm

def _get(ws, r: int, hm: Dict[str,int], key: str):
    c = hm.get(key)
    return ws.cell(r, c).value if c else None

def _set(ws, r: int, hm: Dict[str,int], key: str, value):
    c = hm.get(key)
    if c:
        ws.cell(r, c).value = value

def _iso_now():
    return datetime.now().isoformat(timespec="seconds")

def _parse_iso(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace("Z",""))
    except Exception:
        return None

def _snippet(text: str, n: int = 180) -> str:
    t = (text or "").strip().replace("\n"," ")
    return t[:n]


def run_followups(
    svc,
    outreach_master: Path,
    followup_template_path: Path,
    max_followups_per_run: int = 10,
) -> int:
    wb = load_workbook(outreach_master)
    ws = wb["Outreach"]
    hm = _header_map(ws)

    template = followup_template_path.read_text(encoding="utf-8")

    now = datetime.now()
    sent = 0

    for r in range(2, ws.max_row + 1):
        if sent >= max_followups_per_run:
            break

        status = str(_get(ws, r, hm, "status") or "").strip().upper()
        if status != "SENT":
            continue

        due = _parse_iso(_get(ws, r, hm, "followup_due_at"))
        if not due or due > now:
            continue

        to_email = str(_get(ws, r, hm, "to_email") or "").strip()
        to_name  = str(_get(ws, r, hm, "to_name") or "there").strip()
        title    = str(_get(ws, r, hm, "job_title") or "the role").strip()
        company  = str(_get(ws, r, hm, "company") or "your company").strip()
        thread_id = str(_get(ws, r, hm, "thread_id") or "").strip()

        body = template.format(name=to_name, title=title, company=company)
        subject = f"Follow-up: {title} — Umar"

        # send follow-up in SAME thread if thread exists
        send_email_with_pdf(
            svc,
            to_email=to_email,
            subject=subject,
            body_text=body,
            pdf_path=CV_PDF_PATH,
            thread_id=thread_id or None
        )

        _set(ws, r, hm, "status", "FOLLOWUP_SENT")
        _set(ws, r, hm, "followup_sent_at", _iso_now())
        sent += 1

    wb.save(outreach_master)
    return sent


def run_reply_check(
    svc,
    outreach_master: Path,
    replies_master: Path,
    your_email_hint: str = "umar",   # simple filter for "my own messages"
) -> int:
    """
    Reads thread_id from outreach_master, checks Gmail thread.
    If message is from someone else (not your_email_hint), logs it.
    """
    # load outreach
    owb = load_workbook(outreach_master)
    ows = owb["Outreach"]
    ohm = _header_map(ows)

    # load replies
    rwb = load_workbook(replies_master)
    rws = rwb["Replies"]

    reply_added = 0

    for r in range(2, ows.max_row + 1):
        thread_id = str(_get(ows, r, ohm, "thread_id") or "").strip()
        if not thread_id:
            continue

        status = str(_get(ows, r, ohm, "status") or "").strip().upper()
        if status == "STOP":
            continue

        person_id = str(_get(ows, r, ohm, "person_id") or "").strip()
        to_email = str(_get(ows, r, ohm, "to_email") or "").strip()

        # fetch thread
        thread = svc.users().threads().get(userId="me", id=thread_id).execute()
        messages = thread.get("messages", [])

        for msg in messages:
            headers = {h["name"].lower(): h["value"] for h in msg["payload"]["headers"]}
            from_email = headers.get("from", "")
            subject = headers.get("subject", "")

            # ignore your own sent messages
            if your_email_hint.lower() in from_email.lower():
                continue

            # internalDate
            reply_at = datetime.fromtimestamp(int(msg["internalDate"]) / 1000).isoformat(timespec="seconds")

            # snippet
            snippet = msg.get("snippet", "")

            # Append to replies sheet
            rws.append([reply_at, from_email, subject, _snippet(snippet), thread_id, person_id])

            # Update outreach row
            _set(ows, r, ohm, "status", "REPLIED")
            _set(ows, r, ohm, "last_reply_at", reply_at)
            _set(ows, r, ohm, "last_reply_snippet", _snippet(snippet))
            reply_added += 1

            # only log first external reply per thread per run
            break

    rwb.save(replies_master)
    owb.save(outreach_master)
    return reply_added
